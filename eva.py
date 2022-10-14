# -*- coding: utf-8 -*-
import pandas as pd
import openpyxl as op
import numpy as np
import sys, getopt, datetime

# import pretty_errors

# 命令行接收filepath
def parse_argv():
    simulation_filepath = ""
    observation_filepath = ""
    output_filepath = ""
    # "-sim" 为命令行参数 用于指定nc文件路径, "-obs" 为命令行参数 用于指定观测值文件路径, "-out" 为命令行参数 用于指定输出文件路径
    try:
        opts, args = getopt.getopt(sys.argv[1:], "s:b:o:", ["sim=", "obs=", "out="])
    except getopt.GetoptError:
        print(
            "python a.py -s <simulation_filepath> -b <observation_filepath> -o <output_filepath>"
        )
        sys.exit(2)

    for opt, arg in opts:
        if opt == "-h":
            print(
                "python a.py -s <simulation_filepath> -b <observation_filepath> -o <output_filepath>"
            )
            sys.exit()
        elif opt in ["-s", "--sim"]:
            simulation_filepath = arg
        elif opt in ("-b", "--obs"):
            observation_filepath = arg
        elif opt in ("-o", "--out"):
            output_filepath = arg
    return simulation_filepath, observation_filepath, output_filepath


"""
使用openpyxl 完善标题等数据
"""


def gen_template(output_filepath):
    # 读取表格
    book = op.load_workbook(output_filepath)
    # sheet 名为2020
    sheet = book["2020"]

    # 第2行上面插入3行
    sheet.insert_rows(2, 3)
    # 在第1行上面插入1行
    sheet.insert_rows(1, 1)

    """
    第一行
    """

    # 合并C1-E1
    sheet.merge_cells("C1:E1")
    # 设置内容为T(℃)
    sheet["C1"] = "T(℃)"

    # 合并F1-H1
    sheet.merge_cells("F1:H1")
    # 设置内容为 RH(%S)
    sheet["F1"] = "RH(%S)"

    # 合并I1-K1
    sheet.merge_cells("I1:K1")
    # 设置内容为P(Pa)
    sheet["I1"] = "P(Pa)"

    # 合并L1-N1
    sheet.merge_cells("L1:N1")
    # 设置内容为WS(m/s)
    sheet["L1"] = "WS(m/s)"

    """
    第二行
    """
    # 设置C2\F2\I2\L2内容为“d01模拟浓度”
    sheet["C2"] = "d01模拟浓度"
    sheet["F2"] = "d01模拟浓度"
    sheet["I2"] = "d01模拟浓度"
    sheet["L2"] = "d01模拟浓度"
    # 设置D2\G2\J2\M2内容为“d02模拟浓度”
    sheet["D2"] = "d02模拟浓度"
    sheet["G2"] = "d02模拟浓度"
    sheet["J2"] = "d02模拟浓度"
    sheet["M2"] = "d02模拟浓度"
    # 设置E2\H2\K2\N2内容为“观测值”
    sheet["E2"] = "观测值"
    sheet["H2"] = "观测值"
    sheet["K2"] = "观测值"
    sheet["N2"] = "观测值"

    """
    b列
    """
    # 设置B3 内容为r
    sheet["B3"] = "r"
    # 设置B4 内容为MB
    sheet["B4"] = "MB"
    # 设置B5 内容为NMB
    sheet["B5"] = "NMB"

    """
    给所有单元格设置框线需要遍历全部单元格的xml文件，特别耗费时间，所以注释掉，建议手动设置
    """
    # 给整个表格设置所有框线
    # for row in sheet.iter_rows():
    #     for cell in row:
    #         cell.border = my_border("thin", "thin", "thin", "thin")

    # 保存文件
    book.save(output_filepath)


"""
使用pandas 解析观测数据文件
"""


def parse_observation(observation_filepath):
    # 读取数据
    df = pd.read_csv(observation_filepath, low_memory=False, sep=",", encoding="utf-8")
    # 去除空行
    df.dropna(inplace=True)

    """
    关于时间的提取,需要先排序
    """
    # 去除前3列
    df = df.iloc[:, 3:]
    # 前3列只保留整数位，合并为 格式为 YYYY-MM-DD
    df["date"] = (
        df.iloc[:, :3].astype(int).astype(str).apply(lambda x: "-".join(x), axis=1)
    )
    # 去除前3列
    df = df.iloc[:, 3:]
    # 将date列转化为日期格式
    df["date"] = pd.to_datetime(df["date"])
    # 根据日期排序
    df = df.sort_values(by="date")
    # date分组,Hour列排序
    df = df.groupby("date").apply(lambda x: x.sort_values(by="Hour"))
    # 重置索引
    df = df.reset_index(drop=True)

    # dates放到list中o
    dates_list = df["date"].tolist()
    # 只保留年月日
    dates_list = [date.date() for date in dates_list]

    Hours = df["Hour"]
    # 将时刻转化为字符串
    Hours_list = [str(int(Hours)) for Hours in Hours]

    """
    观测值的提取
    """
    # 读取PSFC、RH、T2、WS列
    PSFC = df["PSFC"]
    T2 = df["T2"]
    RH = df["RH"]
    WS = df["WS"]

    # 将观测值转化为字符串list
    PSFC_list = [str(PSFC) for PSFC in PSFC]
    T2_list = [str(T2) for T2 in T2]
    RH_list = [str(RH) for RH in RH]
    WS_list = [str(WS) for WS in WS]

    return dates_list, Hours_list, PSFC_list, T2_list, RH_list, WS_list


"""
使用pandas 写入观测数据和模拟数据
"""


def data2excle(
    output_filepath,
    dates_list,
    Hours_list,
    PSFC_list,
    T2_list,
    RH_list,
    WS_list,
    d01,
    d02,
):
    # 创建一个空的dataframe
    df = pd.DataFrame()

    # 将各list写入dataframe
    df["日期"] = pd.DataFrame(dates_list)
    df["时刻"] = pd.DataFrame(Hours_list)

    df["d01模拟浓度T2"] = pd.DataFrame(d01[0])
    df["d02模拟浓度T2"] = pd.DataFrame(d02[0])
    df["观测值T2"] = pd.DataFrame(T2_list)

    df["d01模拟浓度RH"] = pd.DataFrame(d01[1])
    df["d02模拟浓度RH"] = pd.DataFrame(d02[1])
    df["观测值RH"] = pd.DataFrame(RH_list)

    df["d01模拟浓度PSFC"] = pd.DataFrame(d01[2])
    df["d02模拟浓度PSFC"] = pd.DataFrame(d02[2])
    df["观测值PSFC"] = pd.DataFrame(PSFC_list)

    df["d01模拟浓度WS"] = pd.DataFrame(d01[3])
    df["d02模拟浓度WS"] = pd.DataFrame(d02[3])
    df["观测值WS"] = pd.DataFrame(WS_list)

    # 将dataframe写入excle文件，sheet为2020
    df.to_excel(output_filepath, sheet_name="2020", index=False)


"""
解析模拟浓度nc文件
"""


def parse_nc(simulation_filepath):
    import netCDF4 as nc

    # 读取nc文件
    nc = nc.Dataset(simulation_filepath, "r")
    # 读取nc文件中的矩阵
    d01 = nc.variables["data_sim"][:][0][0]
    d02 = nc.variables["data_sim"][:][1][0]
    # 转置矩阵
    # 共4组 分别为T、RH、PSFC、WS 的模拟浓度
    d01 = d01.T
    d02 = d02.T
    return np.array(d01), np.array(d02)


"""
计算模拟浓度和观测值的相关系数
"""


def evaluation(d01, d02, T2_list, RH_list, PSFC_list, WS_list):
    T2_list = [float(T2) for T2 in T2_list]
    RH_list = [float(RH) for RH in RH_list]
    PSFC_list = [float(PSFC) for PSFC in PSFC_list]
    WS_list = [float(WS) for WS in WS_list]

    """
    去除nan值
    """
    # T2_list 中 含有 999999.0 的值，需要去除，对应的模拟浓度也需要去除

    """
    去除T2_list中的999999.0
    """
    # 构建新的list
    T2_list_new = []
    d01_T2_new = []
    d02_T2_new = []

    # 循环遍历 T2_list，将不为999999.0的值添加到新的list中
    for i in range(len(T2_list)):
        if T2_list[i] != 999999.0:
            T2_list_new.append(T2_list[i])
            d01_T2_new.append(d01[0][i])
            d02_T2_new.append(d02[0][i])

    """
    去除RH_list中的999999.0
    """
    RH_list_new = []
    d01_RH_new = []
    d02_RH_new = []

    for i in range(len(RH_list)):
        if RH_list[i] != 999999.0:
            RH_list_new.append(RH_list[i])
            d01_RH_new.append(d01[1][i])
            d02_RH_new.append(d02[1][i])

    """
    去除PSFC_list中的999999.0
    """

    PSFC_list_new = []
    d01_PSFC_new = []
    d02_PSFC_new = []

    for i in range(len(PSFC_list)):
        if PSFC_list[i] != 999999.0:
            PSFC_list_new.append(PSFC_list[i])
            d01_PSFC_new.append(d01[2][i])
            d02_PSFC_new.append(d02[2][i])
    """
    去除WS_list中的999999.0
    """
    WS_list_new = []
    d01_WS_new = []
    d02_WS_new = []

    for i in range(len(WS_list)):
        if WS_list[i] != 999999.0:
            WS_list_new.append(WS_list[i])
            d01_WS_new.append(d01[3][i])
            d02_WS_new.append(d02[3][i])

    """
    计算相关系数
    """
    # 计算r
    r_d01_T2 = np.corrcoef(d01_T2_new, T2_list_new)[0][1]
    r_d02_T2 = np.corrcoef(d02_T2_new, T2_list_new)[0][1]
    r_d01_RH = np.corrcoef(d01_RH_new, RH_list_new)[0][1]
    r_d02_RH = np.corrcoef(d02_RH_new, RH_list_new)[0][1]
    r_d01_PSFC = np.corrcoef(d01_PSFC_new, PSFC_list_new)[0][1]
    r_d02_PSFC = np.corrcoef(d02_PSFC_new, PSFC_list_new)[0][1]
    r_d01_WS = np.corrcoef(d01_WS_new, WS_list_new)[0][1]
    r_d02_WS = np.corrcoef(d02_WS_new, WS_list_new)[0][1]

    # 打包r
    r_list = [
        r_d01_T2,
        r_d02_T2,
        r_d01_RH,
        r_d02_RH,
        r_d01_PSFC,
        r_d02_PSFC,
        r_d01_WS,
        r_d02_WS,
    ]

    # 计算差的均值MB
    MB_d01_T2 = np.mean(np.array(d01_T2_new[0]) - np.array(T2_list_new))
    MB_d02_T2 = np.mean(np.array(d02_T2_new[0]) - np.array(T2_list_new))
    MB_d01_RH = np.mean(np.array(d01_RH_new[0]) - np.array(RH_list_new))
    MB_d02_RH = np.mean(np.array(d02_RH_new[0]) - np.array(RH_list_new))
    MB_d01_PSFC = np.mean(np.array(d01_PSFC_new[0]) - np.array(PSFC_list_new))
    MB_d02_PSFC = np.mean(np.array(d02_PSFC_new[0]) - np.array(PSFC_list_new))
    MB_d01_WS = np.mean(np.array(d01_WS_new[0]) - np.array(WS_list_new))
    MB_d02_WS = np.mean(np.array(d02_WS_new[0]) - np.array(WS_list_new))

    # 打包MB
    MB_list = [
        MB_d01_T2,
        MB_d02_T2,
        MB_d01_RH,
        MB_d02_RH,
        MB_d01_PSFC,
        MB_d02_PSFC,
        MB_d01_WS,
        MB_d02_WS,
    ]

    # 计算NMB NMB= MB / (观测值均值) * 100
    NMB_d01_T2 = MB_d01_T2 / np.mean(np.array(T2_list_new)) * 100
    NMB_d02_T2 = MB_d02_T2 / np.mean(np.array(T2_list_new)) * 100
    NMB_d01_RH = MB_d01_RH / np.mean(np.array(RH_list_new)) * 100
    NMB_d02_RH = MB_d02_RH / np.mean(np.array(RH_list_new)) * 100
    NMB_d01_PSFC = MB_d01_PSFC / np.mean(np.array(PSFC_list_new)) * 100
    NMB_d02_PSFC = MB_d02_PSFC / np.mean(np.array(PSFC_list_new)) * 100
    NMB_d01_WS = MB_d01_WS / np.mean(np.array(WS_list_new)) * 100
    NMB_d02_WS = MB_d02_WS / np.mean(np.array(WS_list_new)) * 100

    # 打包NMB
    NMB_list = [
        NMB_d01_T2,
        NMB_d02_T2,
        NMB_d01_RH,
        NMB_d02_RH,
        NMB_d01_PSFC,
        NMB_d02_PSFC,
        NMB_d01_WS,
        NMB_d02_WS,
    ]

    return r_list, MB_list, NMB_list


def evulation2excle(r_list, MB_list, NMB_list, output_filepath):
    book = op.load_workbook(output_filepath)
    sheet = book.active

    # 写入相关系数
    # 分别在C3\D3\F3\G3\I3\J3\L3\M3 位置写入 r_list的数据
    sheet["C3"] = r_list[0]
    sheet["D3"] = r_list[1]
    sheet["F3"] = r_list[2]
    sheet["G3"] = r_list[3]
    sheet["I3"] = r_list[4]
    sheet["J3"] = r_list[5]
    sheet["L3"] = r_list[6]
    sheet["M3"] = r_list[7]

    # 写入MB
    # 分别在C4\D4\F4\G4\I4\J4\L4\M4 位置写入 MB_list的数据
    sheet["C4"] = MB_list[0]
    sheet["D4"] = MB_list[1]
    sheet["F4"] = MB_list[2]
    sheet["G4"] = MB_list[3]
    sheet["I4"] = MB_list[4]
    sheet["J4"] = MB_list[5]
    sheet["L4"] = MB_list[6]
    sheet["M4"] = MB_list[7]

    # 写入NMB
    # 分别在C5\D5\F5\G5\I5\J5\L5\M5 位置写入 NMB_list的数据
    sheet["C5"] = NMB_list[0]
    sheet["D5"] = NMB_list[1]
    sheet["F5"] = NMB_list[2]
    sheet["G5"] = NMB_list[3]
    sheet["I5"] = NMB_list[4]
    sheet["J5"] = NMB_list[5]
    sheet["L5"] = NMB_list[6]
    sheet["M5"] = NMB_list[7]

    book.save(output_filepath)


if __name__ == "__main__":

    # 从终端获取参数
    # simulation_filepath, observation_filepath, output_filepath = parse_argv()

    # parse_observation(observation_filepath)
    simulation_filepath = r"C:\Users\fanyq\Desktop\store\source_code\sim_data.nc"
    observation_filepath = r"C:\Users\fanyq\Desktop\store\source_code\202007ob.csv"
    output_filepath = r"C:\Users\fanyq\Desktop\store\source_code\out.xlsx"

    # 解析模拟数据
    d01, d02 = parse_nc(simulation_filepath)
    # 解析观测数据
    dates_list, Hours_list, PSFC_list, T2_list, RH_list, WS_list = parse_observation(
        observation_filepath
    )

    # 写入观测和模拟的数据
    data2excle(
        output_filepath,
        dates_list,
        Hours_list,
        PSFC_list,
        T2_list,
        RH_list,
        WS_list,
        d01,
        d02,
    )

    # 完善excle文件 耗费时间较长大概4500ms
    gen_template(output_filepath)

    # 计算相关系数
    r_list, MB_list, NMB_list = evaluation(
        d01, d02, T2_list, RH_list, PSFC_list, WS_list
    )

    # 写入计算结果
    evulation2excle(r_list, MB_list, NMB_list, output_filepath)
